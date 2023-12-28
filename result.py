
import pandas 
import openpyxl
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
import time

# Inicializēt WebDriver
service = Service()
driver = webdriver.Chrome(service=service)
wait = WebDriverWait(driver, 10)

words=[]                # angļu valodas vārdi 
Translations = []       # latviešu valodas vārdi
Translations_2 = []     # krievu valodas vārdi

fails = pandas.read_excel("Words.xlsx", sheet_name="Sheet1")
info_list = fails.values.tolist()

for row in info_list:           #paņēm vārdus no "Sheet1" un ievadā masīvā: "words[]"
    words.append(row[0])

words.sort() # sakārto pēc alfabētas secība

##print(words)

#Tulkotāja "URL", kurš tulko latviešu valodā 

url = "https://translate.yandex.com/en/?source_lang=en&target_lang=lv"


for word in words:          #Cikls kurš partilko vārdus uz latviešu valodu
    
    driver.get(url)         # saņēm "URL"

    input_field=driver.find_element(By.XPATH, "//*[@id='fakeArea']")            #atrod lauku kur jāievada vārdu

    time.sleep(2)           #aizkavēšana, lai tīmēkļa lapa domātu, ka mēs esam cilvēks                        
    
    input_field.clear()             #notirā lauku
    input_field.send_keys(word)     #ievieto vārdu laukā
    
    time.sleep(2)

    translated_text = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[contains(@class, 'translation')]"))).text   #nokopē pārtulkoto vārdu ar aizkavēšanu 

    
    print(translated_text)  #izdruka pārtulkotu vārdu, kurš tikko boja iekopēts
    
    Translations.append(translated_text)    #pievieno pārtulkotu vārdu masīvā "Translations[]"


#Tulkotāja "URL", kurš tulko krievu valodā 
    
url_2 = "https://translate.yandex.com/en/?source_lang=en&target_lang=ru"


for word in words:
    
    driver.get(url_2)

    input_field=driver.find_element(By.XPATH, "//*[@id='fakeArea']")

    time.sleep(2)
    
    input_field.clear()
    input_field.send_keys(word)
    
    time.sleep(2)

    translated_text = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[contains(@class, 'translation')]"))).text

    print(translated_text)  #izdruka pārtulkotu vārdu, kurš tikko boja iekopēts
    
    Translations_2.append(translated_text) #pievieno pārtulkotu vārdu masīvā "Translations_2[]"

driver.close() #slēdz tīmekļa lapu

wb = openpyxl.load_workbook('Words.xlsx')   # ielādē Excel darba grāmatu no faila 'Words.xlsx'
ws = wb['result']                           #  iegūst 'result' darba lapu 

bold_font = Font(bold=True)                 #Mainīgais tiek izmantots, lai formatētu tekstu excel uz "bold"


center_aligned_text = Alignment(horizontal='center')        # Mainīgais tiek izmantots, lai izveidot centrētu izlīdzināšanu

#Ierakstā 'A1','B1','C1' virsrakstus
#Maina 'A1','B1','C1' tekstu uz "bold"
#Izveidot centrētu izlīdzināšanu 'A1','B1','C1' laukos
#Maina 'A1','B1','C1' platumu uz 20, 20, 50 px.

ws['A1'] = 'English'          
ws['A1'].font = bold_font
ws['A1'].alignment = center_aligned_text

ws['B1'] = 'Latvian'
ws['B1'].font = bold_font
ws['B1'].alignment = center_aligned_text

ws['C1'] = 'Russian'
ws['C1'].font = bold_font
ws['C1'].alignment = center_aligned_text

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 50


fails = pandas.read_excel("Words.xlsx", sheet_name="Sheet1") #rinda ielasīs datus no Excel faila, kas atrodas faila "Words.xlsx" lapā "Sheet1"
info_list = fails.values.tolist()                            # pārveidos  datus par sarakstu (list) formātā

#Ievieto vārdus no 3 masīviem kolonas veida 
for i in range(len(info_list)):
    ws.append([words[i], Translations[i], Translations_2[i]])

# Saglabāt darbgrāmatu
wb.save('Words.xlsx')

